import base64
import hashlib
import io
import math
import uuid
from datetime import date

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from streamlit_geolocation import streamlit_geolocation

# MÓDULO SUPABASE
from supabase import Client, create_client

# IMPORTAÇÃO DOS MÓDULOS CUSTOMIZADOS
from ui_components import aplicar_estilo_customizado, render_kpi_card

# ==============================================================================
# INICIALIZAÇÃO DA APLICAÇÃO E SESSION STATE
# ==============================================================================

st.set_page_config(
    page_title="Central de Controle - Sondagem",
    page_icon=None,
    layout="wide"
)

# Estilo CSS para ocultar elementos de marca, ícones nativos e menus do Streamlit
hide_streamlit_style = """
    <style>
    /* Oculta o menu de três pontos do Streamlit */
    #MainMenu {visibility: hidden;}
    
    /* Oculta o cabeçalho padrão e a barra superior */
    header {visibility: hidden;}
    .stApp > header {display: none;}
    
    /* Oculta o rodapé e o painel 'Manage app' */
    footer {visibility: hidden;}
    [data-testid="stStatusWidget"] {visibility: hidden;}
    [data-testid="stDecoration"] {display: none;}
    
    /* Oculta ícones SVG internos de menus/navegação do Streamlit */
    [data-testid="stSidebarNav"] svg {display: none;}
    button[title="View fullscreen"] {display: none;}

    /* Ajustes leves para uso em celular */
    @media (max-width: 768px) {
        .block-container {padding: 1rem 0.7rem 4rem 0.7rem;}
        [data-testid="stHorizontalBlock"] {gap: 0.5rem;}
        button {min-height: 44px;}
    }
    </style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False
if "usuario" not in st.session_state:
    st.session_state["usuario"] = ""
if "perfil" not in st.session_state:
    st.session_state["perfil"] = ""
if "sonda_id" not in st.session_state:
    st.session_state["sonda_id"] = None
if "navegacao" not in st.session_state:
    st.session_state["navegacao"] = "Dashboard Geral"


def botao_voltar_dashboard():
    """Botão simples e grande para facilitar a navegação no celular."""
    if st.button("⬅️ Voltar ao Dashboard", use_container_width=True):
        st.session_state["navegacao"] = "Dashboard Geral"
        st.rerun()


# ==============================================================================
# CONEXÃO E CONFIGURAÇÃO DO SUPABASE
# ==============================================================================

SUPABASE_URL = st.secrets.get("SUPABASE_URL", "https://seu-projeto.supabase.co")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "sua-chave-aqui")


@st.cache_resource
def get_supabase_client() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def upload_foto_supabase(file_buffer, nome_arquivo):
    try:
        supabase = get_supabase_client()
        bucket_name = "boletins-fotos"
        ext = nome_arquivo.split(".")[-1]
        caminho_storage = f"boletins/{uuid.uuid4().hex}.{ext}"

        supabase.storage.from_(bucket_name).upload(
            caminho_storage,
            file_buffer.getvalue(),
            file_options={"content-type": f"image/{ext}"},
        )

        return supabase.storage.from_(bucket_name).get_public_url(caminho_storage)
    except Exception as e:
        st.error(f"Erro ao enviar imagem para o Supabase Storage: {e}")
        return None


def salvar_boletim_supabase(dados_boletim):
    try:
        supabase = get_supabase_client()
        return supabase.table("boletim_geologico").insert(dados_boletim).execute()
    except Exception as e:
        st.error(f"Erro ao salvar boletim no Supabase: {e}")
        return None


# ==============================================================================
# BANCO DE DADOS & AUTENTICAÇÃO
# ==============================================================================


def hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()


def get_supabase():
    return get_supabase_client()


def _rows(response):
    return response.data if response and response.data else []


def table_df(table_name, columns="*"):
    """Lê uma tabela do Supabase e devolve DataFrame."""
    try:
        data = _rows(get_supabase().table(table_name).select(columns).execute())
        return pd.DataFrame(data)
    except Exception as e:
        st.error(f"Erro ao consultar a tabela '{table_name}' no Supabase: {e}")
        return pd.DataFrame()


def insert_row(table_name, data):
    return get_supabase().table(table_name).insert(data).execute()


def update_row(table_name, data, column, value):
    return get_supabase().table(table_name).update(data).eq(column, value).execute()


def delete_row(table_name, column, value):
    return get_supabase().table(table_name).delete().eq(column, value).execute()


def criar_tabela_usuarios():
    """
    O banco agora fica 100% no Supabase.
    As tabelas devem ser criadas uma única vez usando o arquivo
    supabase_schema.sql que acompanha este projeto.
    """
    try:
        supabase = get_supabase()
        resposta = supabase.table("usuarios").select("id").limit(1).execute()
        if not resposta.data:
            try:
                supabase.table("usuarios").insert({
                    "usuario": "admin",
                    "senha": hash_senha("admin123"),
                    "perfil": "Admin",
                    "sonda_id": None,
                }).execute()
            except Exception as e:
                # Se o admin já existir, não interrompe a inicialização.
                if "duplicate" not in str(e).lower() and "unique" not in str(e).lower():
                    st.warning(f"Não foi possível criar o usuário admin automaticamente: {e}")
    except Exception as e:
        st.error(
            "Não foi possível acessar o Supabase. Verifique SUPABASE_URL, "
            f"SUPABASE_KEY e se as tabelas foram criadas. Detalhe: {e}"
        )


def verificar_login(usuario, senha):
    try:
        resultado = (
            get_supabase()
            .table("usuarios")
            .select("perfil, sonda_id")
            .eq("usuario", usuario)
            .eq("senha", hash_senha(senha))
            .limit(1)
            .execute()
        )
        dados = _rows(resultado)
        return (dados[0]["perfil"], dados[0].get("sonda_id")) if dados else None
    except Exception as e:
        st.error(f"Erro ao verificar login no Supabase: {e}")
        return None


def obter_dados_dashboard(perfil_usuario, user_sonda_id):
    df_sondas = table_df("sondas")
    df_prod = table_df("producao_diaria")
    df_geo = table_df("boletim_geologico")

    if df_prod.empty:
        df_prod = pd.DataFrame(columns=[
            "id", "data", "sonda_id", "furo_id", "prof_inicial",
            "prof_final", "horas_trabalhadas", "horas_paradas", "motivo_parada"
        ])
    if df_sondas.empty:
        df_sondas = pd.DataFrame(columns=["id", "codigo", "equipe", "projeto", "status"])
    if df_geo.empty:
        df_geo = pd.DataFrame(columns=[
            "id", "furo_id", "de_m", "ate_m", "recuperacao_m", "rqd_m",
            "litologia", "descricao_geologica", "n_amostra", "observacoes", "foto_url"
        ])

    if perfil_usuario != "Admin":
        df_sondas = df_sondas[df_sondas["id"] == user_sonda_id]
        df_prod = df_prod[df_prod["sonda_id"] == user_sonda_id]

        df_furos = table_df("furos")
        if not df_furos.empty:
            df_furos = df_furos[df_furos["sonda_id"] == user_sonda_id]
            ids_furos = set(df_furos["id"].tolist())
            df_geo = df_geo[df_geo["furo_id"].isin(ids_furos)]
        else:
            df_geo = df_geo.iloc[0:0]

    if not df_prod.empty and not df_sondas.empty:
        df_prod = df_prod.merge(
            df_sondas[["id", "codigo"]],
            left_on="sonda_id", right_on="id", how="left", suffixes=("", "_sonda")
        )
        df_prod.rename(columns={"codigo": "sonda_codigo"}, inplace=True)

    if not df_geo.empty:
        df_geo = df_geo.copy()
        df_geo["avanco_m"] = df_geo["ate_m"] - df_geo["de_m"]
        df_geo["recuperacao_pct"] = (
            df_geo["recuperacao_m"] / df_geo["avanco_m"].replace(0, pd.NA) * 100
        ).round(1)
        df_geo["rqd_pct"] = (
            df_geo["rqd_m"] / df_geo["avanco_m"].replace(0, pd.NA) * 100
        ).round(1)

    return df_sondas, df_prod, df_geo

def tela_login():
    criar_tabela_usuarios()

    if not st.session_state.get("autenticado", False):
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.subheader("Acesso ao Sistema")
            usuario = st.text_input("Usuário")
            senha = st.text_input("Senha", type="password")

            if st.button("Entrar", type="primary", use_container_width=True):
                login_info = verificar_login(usuario, senha)
                if login_info:
                    st.session_state["autenticado"] = True
                    st.session_state["usuario"] = usuario
                    st.session_state["perfil"] = login_info[0]
                    st.session_state["sonda_id"] = login_info[1]
                    st.success("Login realizado com sucesso!")
                    st.rerun()
                else:
                    st.error("Usuário ou senha incorretos.")
        return False
    return True


def botao_logout():
    st.sidebar.markdown(
        f"**{st.session_state.get('usuario', '')}** ({st.session_state.get('perfil', '')})"
    )
    if st.sidebar.button("Sair / Logout"):
        st.session_state["autenticado"] = False
        st.session_state["usuario"] = ""
        st.session_state["perfil"] = ""
        st.session_state["sonda_id"] = None
        st.rerun()


# ==============================================================================
# HELPER DE EXPORTAÇÃO EXCEL
# ==============================================================================


def gerar_dashboard_excel_completo(perfil_usuario, user_sonda_id):
    """Gera um relatório Excel completo, colorido e atualizado com os dados do SQLite.
    Mantém o aplicativo original intacto e apenas amplia a exportação.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Tema visual
    # ------------------------------------------------------------------
    C = {
        "navy": "17365D", "blue": "1F4E79", "blue2": "5B9BD5",
        "cyan": "00B0F0", "green": "70AD47", "green2": "C6E0B4",
        "yellow": "FFC000", "orange": "F4B183", "red": "C00000",
        "red2": "F4CCCC", "gray": "D9E1F2", "light": "F3F6FA",
        "white": "FFFFFF", "dark": "1F2937", "border": "B7C9D6",
        "purple": "8064A2", "purple2": "E4DFEC"
    }
    fill_title = PatternFill("solid", fgColor=C["navy"])
    fill_header = PatternFill("solid", fgColor=C["blue"])
    fill_sub = PatternFill("solid", fgColor=C["gray"])
    fill_card = PatternFill("solid", fgColor=C["light"])
    fill_green = PatternFill("solid", fgColor=C["green2"])
    fill_yellow = PatternFill("solid", fgColor="FFF2CC")
    fill_red = PatternFill("solid", fgColor=C["red2"])
    fill_purple = PatternFill("solid", fgColor=C["purple2"])
    thin = Side(style="thin", color=C["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="Calibri", size=16, bold=True, color=C["white"])
    header_font = Font(name="Calibri", size=10, bold=True, color=C["white"])
    section_font = Font(name="Calibri", size=12, bold=True, color=C["navy"])
    normal_font = Font(name="Calibri", size=10, color=C["dark"])
    small_font = Font(name="Calibri", size=9, color="666666")
    card_label_font = Font(name="Calibri", size=9, bold=True, color="666666")
    card_value_font = Font(name="Calibri", size=16, bold=True, color=C["blue"])

    # ------------------------------------------------------------------
    # Leitura dos dados do banco local original
    # ------------------------------------------------------------------
    conn = get_connection()
    try:
        if perfil_usuario == "Admin":
            df_sondas = pd.read_sql_query("SELECT * FROM sondas ORDER BY id", conn)
            df_prod = pd.read_sql_query(
                """SELECT p.id, p.data, p.sonda_id, s.codigo AS sonda_codigo, s.equipe,
                          s.projeto, p.furo_id, p.prof_inicial, p.prof_final,
                          (p.prof_final - p.prof_inicial) AS avanco,
                          p.horas_trabalhadas, p.horas_paradas, p.motivo_parada
                   FROM producao_diaria p
                   LEFT JOIN sondas s ON p.sonda_id = s.id
                   ORDER BY p.data DESC, p.id DESC""", conn)
            df_furos = pd.read_sql_query(
                """SELECT f.*, s.codigo AS sonda_codigo, s.equipe, s.projeto
                   FROM furos f LEFT JOIN sondas s ON f.sonda_id=s.id
                   ORDER BY f.id""", conn)
            df_geo = pd.read_sql_query(
                """SELECT bg.id, bg.furo_id, f.sonda_id, s.codigo AS sonda_codigo,
                          bg.de_m, bg.ate_m, (bg.ate_m-bg.de_m) AS avanco_m,
                          bg.recuperacao_m,
                          CASE WHEN (bg.ate_m-bg.de_m) > 0 THEN
                            ROUND((bg.recuperacao_m/(bg.ate_m-bg.de_m))*100,1) ELSE 0 END AS recuperacao_pct,
                          bg.rqd_m,
                          CASE WHEN (bg.ate_m-bg.de_m) > 0 THEN
                            ROUND((bg.rqd_m/(bg.ate_m-bg.de_m))*100,1) ELSE 0 END AS rqd_pct,
                          bg.litologia, bg.n_amostra, bg.descricao_geologica,
                          bg.observacoes, bg.foto_url
                   FROM boletim_geologico bg
                   LEFT JOIN furos f ON bg.furo_id=f.id
                   LEFT JOIN sondas s ON f.sonda_id=s.id
                   ORDER BY bg.furo_id, bg.de_m""", conn)
            df_users = pd.read_sql_query(
                """SELECT u.id, u.usuario, u.perfil, u.sonda_id,
                          s.codigo AS sonda_vinculada
                   FROM usuarios u LEFT JOIN sondas s ON u.sonda_id=s.id
                   ORDER BY u.id""", conn)
        else:
            df_sondas = pd.read_sql_query("SELECT * FROM sondas WHERE id = ?", conn, params=(user_sonda_id,))
            df_prod = pd.read_sql_query(
                """SELECT p.id, p.data, p.sonda_id, s.codigo AS sonda_codigo, s.equipe,
                          s.projeto, p.furo_id, p.prof_inicial, p.prof_final,
                          (p.prof_final-p.prof_inicial) AS avanco,
                          p.horas_trabalhadas, p.horas_paradas, p.motivo_parada
                   FROM producao_diaria p LEFT JOIN sondas s ON p.sonda_id=s.id
                   WHERE p.sonda_id=? ORDER BY p.data DESC, p.id DESC""", conn, params=(user_sonda_id,))
            df_furos = pd.read_sql_query(
                """SELECT f.*, s.codigo AS sonda_codigo, s.equipe, s.projeto
                   FROM furos f LEFT JOIN sondas s ON f.sonda_id=s.id
                   WHERE f.sonda_id=? ORDER BY f.id""", conn, params=(user_sonda_id,))
            df_geo = pd.read_sql_query(
                """SELECT bg.id, bg.furo_id, f.sonda_id, s.codigo AS sonda_codigo,
                          bg.de_m, bg.ate_m, (bg.ate_m-bg.de_m) AS avanco_m,
                          bg.recuperacao_m,
                          CASE WHEN (bg.ate_m-bg.de_m)>0 THEN ROUND((bg.recuperacao_m/(bg.ate_m-bg.de_m))*100,1) ELSE 0 END AS recuperacao_pct,
                          bg.rqd_m,
                          CASE WHEN (bg.ate_m-bg.de_m)>0 THEN ROUND((bg.rqd_m/(bg.ate_m-bg.de_m))*100,1) ELSE 0 END AS rqd_pct,
                          bg.litologia, bg.n_amostra, bg.descricao_geologica,
                          bg.observacoes, bg.foto_url
                   FROM boletim_geologico bg
                   JOIN furos f ON bg.furo_id=f.id
                   LEFT JOIN sondas s ON f.sonda_id=s.id
                   WHERE f.sonda_id=? ORDER BY bg.furo_id, bg.de_m""", conn, params=(user_sonda_id,))
            df_users = pd.DataFrame()
    finally:
        conn.close()

    # Garantir colunas e tipos consistentes
    for df in (df_sondas, df_prod, df_furos, df_geo, df_users):
        if df is not None:
            for col in df.columns:
                if str(df[col].dtype).startswith("datetime"):
                    df[col] = df[col].dt.strftime("%d/%m/%Y")
    if not df_prod.empty:
        df_prod["data"] = pd.to_datetime(df_prod["data"], errors="coerce")
    if not df_geo.empty:
        for c in ["de_m", "ate_m", "avanco_m", "recuperacao_m", "recuperacao_pct", "rqd_m", "rqd_pct"]:
            if c in df_geo.columns:
                df_geo[c] = pd.to_numeric(df_geo[c], errors="coerce").fillna(0)

    total_sondas = len(df_sondas)
    sondas_op = int((df_sondas["status"] == "Operando").sum()) if not df_sondas.empty else 0
    sondas_par = int((df_sondas["status"] == "Parada").sum()) if not df_sondas.empty else 0
    sondas_manut = int((df_sondas["status"] == "Manutenção").sum()) if not df_sondas.empty else 0
    total_metros = float(df_prod["avanco"].sum()) if not df_prod.empty else 0.0
    hoje = pd.Timestamp(date.today())
    metros_hoje = float(df_prod.loc[df_prod["data"].dt.normalize() == hoje, "avanco"].sum()) if not df_prod.empty else 0.0
    horas_trab = float(df_prod["horas_trabalhadas"].sum()) if not df_prod.empty else 0.0
    horas_par = float(df_prod["horas_paradas"].sum()) if not df_prod.empty else 0.0
    eficiencia = (horas_trab / (horas_trab + horas_par) * 100) if (horas_trab + horas_par) else 0.0
    media_dia = float(df_prod.groupby(df_prod["data"].dt.date)["avanco"].sum().mean()) if not df_prod.empty else 0.0
    total_furos = len(df_furos)
    furos_concluidos = int((df_furos["situacao"] == "Concluído").sum()) if not df_furos.empty else 0
    total_intervalos = len(df_geo)
    rec_media = float(df_geo["recuperacao_pct"].mean()) if not df_geo.empty else 0.0
    rqd_medio = float(df_geo["rqd_pct"].mean()) if not df_geo.empty else 0.0

    def style_title(ws, title, end_col=10):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        c = ws.cell(1, 1, title)
        c.fill = fill_title; c.font = title_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34
        ws.freeze_panes = "A3"

    def write_df(ws, df, start_row=3, table_name=None):
        if df is None:
            df = pd.DataFrame()
        cols = list(df.columns)
        for j, col in enumerate(cols, 1):
            cell = ws.cell(start_row, j, str(col).replace("_", " ").title())
            cell.fill = fill_header; cell.font = header_font; cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, (_, row) in enumerate(df.iterrows(), start_row + 1):
            for j, val in enumerate(row.tolist(), 1):
                if pd.isna(val): val = ""
                if isinstance(val, pd.Timestamp): val = val.to_pydatetime()
                cell = ws.cell(i, j, val)
                cell.font = normal_font; cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if (i - start_row) % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F8FBFD")
        if cols and len(df) > 0:
            ref = f"A{start_row}:{get_column_letter(len(cols))}{start_row+len(df)}"
            if table_name:
                from openpyxl.worksheet.table import Table, TableStyleInfo
                tab = Table(displayName=table_name, ref=ref)
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
                ws.add_table(tab)
            else:
                ws.auto_filter.ref = ref
        for col_idx in range(1, len(cols)+1):
            letter = get_column_letter(col_idx)
            max_len = max([len(str(ws.cell(r, col_idx).value or "")) for r in range(start_row, min(ws.max_row, start_row+150)+1)] + [10])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 38)
        return start_row + len(df) + 1

    def add_card(ws, label, value, start_col, fill=fill_card):
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col+1)
        ws.merge_cells(start_row=4, start_column=start_col, end_row=5, end_column=start_col+1)
        a = ws.cell(3, start_col, label); a.fill = fill; a.font = card_label_font; a.alignment = Alignment(horizontal="center")
        b = ws.cell(4, start_col, value); b.fill = fill; b.font = card_value_font; b.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=3, max_row=5, min_col=start_col, max_col=start_col+1):
            for cell in row: cell.border = border

    # ------------------------------------------------------------------
    # DASHBOARD EXECUTIVO
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Dashboard Executivo"
    style_title(ws, "DASHBOARD EXECUTIVO — CONTROLE INTEGRADO DE SONDAGEM", 10)
    add_card(ws, "Sondas Operando", f"{sondas_op}/{total_sondas}", 1, fill_green)
    add_card(ws, "Metragem Hoje", f"{metros_hoje:.1f} m", 4)
    add_card(ws, "Metragem Acumulada", f"{total_metros:.1f} m", 7, fill_purple)
    add_card(ws, "Eficiência", f"{eficiencia:.1f}%", 9, fill_yellow)
    ws["A7"] = "Resumo operacional"; ws["A7"].font = section_font
    resumo = pd.DataFrame([
        ["Sondas Operando", sondas_op], ["Sondas Paradas", sondas_par], ["Sondas em Manutenção", sondas_manut],
        ["Furos cadastrados", total_furos], ["Furos concluídos", furos_concluidos],
        ["Intervalos geológicos", total_intervalos], ["Recuperação média", rec_media/100], ["RQD médio", rqd_medio/100],
        ["Horas trabalhadas", horas_trab], ["Horas paradas", horas_par]
    ], columns=["Indicador", "Valor"])
    write_df(ws, resumo, 8, "ResumoOperacional")
    for r in range(9, 19):
        if ws.cell(r,1).value in ["Recuperação média", "RQD médio"]: ws.cell(r,2).number_format = "0.0%"
        elif ws.cell(r,1).value in ["Horas trabalhadas", "Horas paradas"]: ws.cell(r,2).number_format = "0.0"
    # gráfico status
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    status_ws = wb.create_sheet("Dados Gráficos")
    status_ws.sheet_state = "hidden"
    status_ws.append(["Status", "Quantidade"])
    for stt, qtd in [("Operando", sondas_op), ("Parada", sondas_par), ("Manutenção", sondas_manut)]: status_ws.append([stt, qtd])
    pie = PieChart(); pie.title = "Status das Sondas"; pie.height = 7; pie.width = 11
    pie.add_data(Reference(status_ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    pie.set_categories(Reference(status_ws, min_col=1, min_row=2, max_row=4)); ws.add_chart(pie, "D8")
    ws.column_dimensions["A"].width = 27; ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 16; ws.column_dimensions["G"].width = 18
    ws.column_dimensions["I"].width = 18
    ws["A20"] = f"Relatório gerado automaticamente em {date.today().strftime('%d/%m/%Y')} | Perfil: {perfil_usuario}"
    ws["A20"].font = small_font

    # ------------------------------------------------------------------
    # DASHBOARD DE PRODUÇÃO
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Dashboard Produção")
    style_title(ws, "DASHBOARD DE PRODUÇÃO", 12)
    add_card(ws, "Avanço Total", f"{total_metros:.1f} m", 1)
    add_card(ws, "Média por Dia", f"{media_dia:.1f} m", 4, fill_green)
    add_card(ws, "Horas Trabalhadas", f"{horas_trab:.1f} h", 7, fill_purple)
    add_card(ws, "Horas Paradas", f"{horas_par:.1f} h", 10, fill_yellow)
    ws["A7"] = "Produção por Sonda"; ws["A7"].font = section_font
    prod_sonda = (df_prod.groupby("sonda_codigo", dropna=False)["avanco"].sum().reset_index() if not df_prod.empty else pd.DataFrame(columns=["sonda_codigo","avanco"]))
    prod_sonda.columns = ["Sonda", "Avanço (m)"]
    write_df(ws, prod_sonda, 8, "ProducaoPorSonda")
    if not prod_sonda.empty:
        chart = BarChart(); chart.type = "bar"; chart.style = 10; chart.title = "Avanço acumulado por sonda"; chart.y_axis.title = "Sonda"; chart.x_axis.title = "Metros"
        chart.add_data(Reference(ws, min_col=2, min_row=8, max_row=8+len(prod_sonda)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=9, max_row=8+len(prod_sonda))); chart.height=7; chart.width=12; ws.add_chart(chart, "D8")
    ws["A22"] = "Produção diária"; ws["A22"].font = section_font
    if not df_prod.empty:
        diario = df_prod.groupby(df_prod["data"].dt.date)["avanco"].sum().reset_index()
        diario.columns = ["Data", "Avanço (m)"]
    else:
        diario = pd.DataFrame(columns=["Data", "Avanço (m)"])
    write_df(ws, diario, 23, "ProducaoDiaria")
    if not diario.empty:
        chart2 = LineChart(); chart2.title = "Evolução diária do avanço"; chart2.y_axis.title = "Metros"; chart2.x_axis.title = "Data"; chart2.height=7; chart2.width=12
        chart2.add_data(Reference(ws, min_col=2, min_row=23, max_row=23+len(diario)), titles_from_data=True)
        chart2.set_categories(Reference(ws, min_col=1, min_row=24, max_row=23+len(diario))); ws.add_chart(chart2, "D22")

    # ------------------------------------------------------------------
    # DASHBOARD GEOLÓGICO
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Dashboard Geológico")
    style_title(ws, "DASHBOARD GEOLÓGICO", 12)
    add_card(ws, "Intervalos", str(total_intervalos), 1)
    add_card(ws, "Recuperação Média", f"{rec_media:.1f}%", 4, fill_green)
    add_card(ws, "RQD Médio", f"{rqd_medio:.1f}%", 7, fill_purple)
    add_card(ws, "Furos", str(total_furos), 10, fill_yellow)
    ws["A7"] = "Litologias registradas"; ws["A7"].font = section_font
    if not df_geo.empty and "litologia" in df_geo:
        lit = df_geo["litologia"].fillna("Não informado").astype(str).value_counts().reset_index()
        lit.columns = ["Litologia", "Intervalos"]
    else:
        lit = pd.DataFrame(columns=["Litologia", "Intervalos"])
    write_df(ws, lit, 8, "Litologias")
    if not lit.empty:
        chart = BarChart(); chart.title = "Distribuição por litologia"; chart.height=7; chart.width=12
        chart.add_data(Reference(ws, min_col=2, min_row=8, max_row=8+len(lit)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=9, max_row=8+len(lit))); ws.add_chart(chart, "D8")
    ws["A22"] = "Indicadores por furo"; ws["A22"].font = section_font
    if not df_geo.empty:
        geo_furo = df_geo.groupby("furo_id").agg(Avanço_m=("avanco_m","sum"), Recuperação_m=("recuperacao_m","sum"), RQD_m=("rqd_m","sum"), Recuperação_pct=("recuperacao_pct","mean"), RQD_pct=("rqd_pct","mean")).reset_index()
    else:
        geo_furo = pd.DataFrame(columns=["furo_id","Avanço_m","Recuperação_m","RQD_m","Recuperação_pct","RQD_pct"])
    write_df(ws, geo_furo, 23, "IndicadoresGeologicos")

    # ------------------------------------------------------------------
    # FICHAS COMPLETAS DE DADOS
    # ------------------------------------------------------------------
    datasets = [
        ("Sondas", df_sondas, "TabelaSondas"),
        ("Produção Diária", df_prod, "TabelaProducao"),
        ("Furos", df_furos, "TabelaFuros"),
        ("Boletim Geológico", df_geo, "TabelaGeologia"),
    ]
    if perfil_usuario == "Admin": datasets.append(("Usuários", df_users, "TabelaUsuarios"))
    for name, df, table_name in datasets:
        ws = wb.create_sheet(name)
        style_title(ws, f"BASE DE DADOS — {name.upper()}", max(6, min(len(df.columns) if not df.empty else 6, 18)))
        write_df(ws, df, 3, table_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        # formatos úteis
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if isinstance(cell.value, (float, int)) and cell.value is not None:
                    if "pct" in str(ws.cell(3, cell.column).value).lower(): cell.number_format = "0.0"
                    elif "data" in str(ws.cell(3, cell.column).value).lower() and hasattr(cell.value, 'strftime'): cell.number_format = "dd/mm/yyyy"
                    else: cell.number_format = '#,##0.00'

    # links rápidos nos dashboards
    for sheet_name, cell, label in [("Dashboard Executivo","A22","Abrir base de Sondas"), ("Dashboard Executivo","A23","Abrir Produção Diária"), ("Dashboard Executivo","A24","Abrir Boletim Geológico")]:
        wsx = wb[sheet_name]
        wsx[cell] = label
        target = {"Abrir base de Sondas":"Sondas", "Abrir Produção Diária":"Produção Diária", "Abrir Boletim Geológico":"Boletim Geológico"}[label]
        wsx[cell].hyperlink = f"#'{target}'!A1"; wsx[cell].style = "Hyperlink"

    # Abas de dashboard primeiro
    order = ["Dashboard Executivo", "Dashboard Produção", "Dashboard Geológico", "Sondas", "Produção Diária", "Furos", "Boletim Geológico"]
    if perfil_usuario == "Admin": order.append("Usuários")
    order.append("Dados Gráficos")
    wb._sheets = [wb[x] for x in order if x in wb.sheetnames]

    wb.properties.title = "Relatório Completo de Sondagem"
    wb.properties.subject = "Dashboard e base operacional"
    wb.properties.creator = "Central de Controle de Sondagem"
    wb.save(output)
    output.seek(0)
    return output


# ==============================================================================
# INICIALIZAÇÃO DA INTERFACE
# ==============================================================================

aplicar_estilo_customizado()

if not tela_login():
    st.stop()

perfil_atual = st.session_state.get("perfil", "")
sonda_id_atual = st.session_state.get("sonda_id", None)

# ==============================================================================
# SIDEBAR E NAVEGAÇÃO POR PERFIL
# ==============================================================================

st.sidebar.title("CENTRAL DE CONTROLE")
botao_logout()
st.sidebar.markdown("---")

# Definição das opções permitidas baseadas no tipo de usuário
if perfil_atual == "Admin":
    opcoes_menu = [
        "Dashboard Geral",
        "Cadastro de Sondas",
        "Apontamento Diário",
        "Controle de Furos",
        "Boletim Geológico",
        "Gestão de Usuários",
    ]
else:
    opcoes_menu = [
        "Dashboard Geral",
        "Apontamento Diário",
        "Controle de Furos",
        "Boletim Geológico",
    ]

opcao = st.sidebar.radio("Navegação", opcoes_menu, key="navegacao")

excel_mestre = gerar_dashboard_excel_completo(perfil_atual, sonda_id_atual)
st.sidebar.markdown("---")
st.sidebar.download_button(
    label="Exportar Relatório",
    data=excel_mestre,
    file_name=f"Relatorio_Sondagem_{date.today()}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)

# ==============================================================================
# COMPONENTES DE TELA DA APLICAÇÃO
# ==============================================================================

# ==============================================================================
# COMPONENTES DE TELA DA APLICAÇÃO
# ==============================================================================

# ------------------------------------------------------------------------------
# 1. DASHBOARD GERAL (ILUMINADO E DESTAQUE DE SONDAS)
# ------------------------------------------------------------------------------
if opcao == "Dashboard Geral":
    st.title("⚡ Painel Geral de Operações & Sondas")
    st.markdown("---")

    df_sondas, df_prod, _ = obter_dados_dashboard(perfil_atual, sonda_id_atual)

    sondas_total = len(df_sondas)
    sondas_op = len(df_sondas[df_sondas["status"] == "Operando"])
    sondas_par = len(df_sondas[df_sondas["status"] == "Parada"])
    sondas_manut = len(df_sondas[df_sondas["status"] == "Manutenção"])

    if not df_prod.empty:
        df_hoje = df_prod[df_prod["data"] == str(date.today())]
        metros_hoje = df_hoje["prof_final"].sum() - df_hoje["prof_inicial"].sum()
        metros_acumulados = (df_prod["prof_final"] - df_prod["prof_inicial"]).sum()
        hrs_trab = df_prod["horas_trabalhadas"].sum()
        hrs_par = df_prod["horas_paradas"].sum()
        eficiencia = ((hrs_trab / (hrs_trab + hrs_par) * 100) if (hrs_trab + hrs_par) > 0 else 0)
    else:
        metros_hoje, metros_acumulados, eficiencia = 0.0, 0.0, 0.0

    # Estilo CSS para Cards Iluminados de KPIs e Sondas
    st.markdown(
        """
        <style>
        .card-destaque {
            background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
            border-left: 5px solid #00f2fe;
            border-radius: 10px;
            padding: 18px;
            margin-bottom: 15px;
            box-shadow: 0 4px 15px rgba(0, 242, 254, 0.15);
            color: #ffffff;
        }
        .card-sonda-op {
            background: rgba(16, 185, 129, 0.1);
            border: 1px solid #10b981;
            border-radius: 8px;
            padding: 12px;
            margin-bottom: 10px;
        }
        .card-sonda-par {
            background: rgba(239, 68, 68, 0.1);
            border: 1px solid #ef4444;
            border-radius: 8px;
            padding: 12px;
            margin-bottom: 10px;
        }
        .card-sonda-manut {
            background: rgba(245, 158, 11, 0.1);
            border: 1px solid #f59e0b;
            border-radius: 8px;
            padding: 12px;
            margin-bottom: 10px;
        }
        .status-badge {
            font-weight: bold;
            padding: 3px 8px;
            border-radius: 4px;
            font-size: 12px;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    # Indicadores KPI com Iluminação de Destaque
    c1, c2, c3, c4 = st.columns(4)
    with c1: 
        st.markdown(f"""
            <div class="card-destaque">
                <span style="font-size: 13px; color: #94a3b8;">🟢 SONDAS ATIVAS</span>
                <h2 style="margin: 5px 0 0 0; color: #4ade80;">{sondas_op} <span style="font-size:16px; color:#cbd5e1;">/ {sondas_total}</span></h2>
            </div>
        """, unsafe_allow_html=True)
    with c2: 
        st.markdown(f"""
            <div class="card-destaque">
                <span style="font-size: 13px; color: #94a3b8;">🎯 PRODUÇÃO HOJE</span>
                <h2 style="margin: 5px 0 0 0; color: #38bdf8;">{metros_hoje:.1f} m</h2>
            </div>
        """, unsafe_allow_html=True)
    with c3: 
        st.markdown(f"""
            <div class="card-destaque">
                <span style="font-size: 13px; color: #94a3b8;">📐 TOTAL ACUMULADO</span>
                <h2 style="margin: 5px 0 0 0; color: #818cf8;">{metros_acumulados:.1f} m</h2>
            </div>
        """, unsafe_allow_html=True)
    with c4: 
        st.markdown(f"""
            <div class="card-destaque">
                <span style="font-size: 13px; color: #94a3b8;">⚡ EFICIÊNCIA OPERACIONAL</span>
                <h2 style="margin: 5px 0 0 0; color: #f43f5e;">{eficiencia:.0f}%</h2>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("### 🚜 Monitoramento Individual de Sondas")

    if not df_sondas.empty:
        grid_cols = st.columns(3)
        for idx, row in df_sondas.iterrows():
            col_idx = idx % 3
            status = row["status"]

            if status == "Operando":
                css_class = "card-sonda-op"
                cor_status = "#10b981"
                icone = "🟢"
            elif status == "Parada":
                css_class = "card-sonda-par"
                cor_status = "#ef4444"
                icone = "🔴"
            else:
                css_class = "card-sonda-manut"
                cor_status = "#f59e0b"
                icone = "🟡"

            # Busca metragem total produzida por essa sonda específica
            if not df_prod.empty:
                prod_sonda = df_prod[df_prod["sonda_id"] == row["id"]]
                m_sonda = (prod_sonda["prof_final"] - prod_sonda["prof_inicial"]).sum() if not prod_sonda.empty else 0.0
            else:
                m_sonda = 0.0

            with grid_cols[col_idx]:
                st.markdown(
                    f"""
                    <div class="{css_class}">
                        <div style="display:flex; justify-content:space-between; align-items:center;">
                            <h3 style="margin:0; font-size:18px;">Sonda {row['codigo']}</h3>
                            <span class="status-badge" style="background:{cor_status}22; color:{cor_status}; border: 1px solid {cor_status};">
                                {icone} {status.upper()}
                            </span>
                        </div>
                        <p style="margin:5px 0; font-size:13px; color:#cbd5e1;"><b>Equipe:</b> {row['equipe']}</p>
                        <p style="margin:2px 0; font-size:13px; color:#cbd5e1;"><b>Projeto:</b> {row['projeto']}</p>
                        <hr style="margin:8px 0; border:0; border-top:1px solid rgba(255,255,255,0.1);">
                        <div style="display:flex; justify-content:space-between; align-items:center;">
                            <span style="font-size:12px; color:#94a3b8;">Avanço Acumulado:</span>
                            <strong style="font-size:15px; color:#38bdf8;">{m_sonda:.1f} m</strong>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
    else:
        st.info("Nenhuma sonda disponível no momento.")

# ------------------------------------------------------------------------------
# 2. GESTÃO DE SONDAS (EXCLUSIVO ADMIN)
# ------------------------------------------------------------------------------
elif opcao == "Cadastro de Sondas" and perfil_atual == "Admin":
    botao_voltar_dashboard()
    st.title("Gestão Central de Sondas")
    st.markdown("---")

    df_sondas = table_df("sondas")

    tab_lista, tab_novo = st.tabs(["Sondas Cadastradas", "Nova Sonda"])

    with tab_lista:
        if not df_sondas.empty:
            st.dataframe(df_sondas, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("🗑️ Excluir Sonda")
            
            opcoes_sonda = [f"{row['id']} - {row['codigo']} ({row['equipe']})" for _, row in df_sondas.iterrows()]
            sonda_excluir_str = st.selectbox("Selecione a Sonda para excluir:", opcoes_sonda)
            sonda_id_excluir = int(sonda_excluir_str.split(" - ")[0])

            if st.button("Excluir Sonda Selecionada", type="primary"):
                delete_row("sondas", "id", sonda_id_excluir)
                st.success("Sonda excluída com sucesso!")
                st.rerun()
        else:
            st.info("Nenhuma sonda cadastrada.")

    with tab_novo:
        with st.form("form_sonda", clear_on_submit=True):
            codigo = st.text_input("Código da Sonda (ex: SD-01)")
            equipe = st.text_input("Equipe Responsável")
            projeto = st.text_input("Projeto / Frente")
            status = st.selectbox("Status", ["Operando", "Parada", "Manutenção"])
            if st.form_submit_button("Cadastrar", type="primary"):
                if codigo and equipe and projeto:
                    try:
                        insert_row("sondas", {
                            "codigo": codigo,
                            "equipe": equipe,
                            "projeto": projeto,
                            "status": status,
                        })
                        st.success("Sonda cadastrada com sucesso!")
                        st.rerun()
                    except Exception as e:
                        if "duplicate" in str(e).lower() or "unique" in str(e).lower():
                            st.error("Sonda já cadastrada.")
                        else:
                            st.error(f"Erro ao cadastrar sonda: {e}")

# ------------------------------------------------------------------------------
# 3. APONTAMENTO DIÁRIO
# ------------------------------------------------------------------------------
elif opcao == "Apontamento Diário":
    botao_voltar_dashboard()
    st.title("Apontamento Diário de Produção")
    st.markdown("---")

    df_sondas = table_df("sondas")[["id", "codigo"]] if not table_df("sondas").empty else pd.DataFrame(columns=["id", "codigo"])
    df_furos = table_df("furos")
    df_prod = table_df("producao_diaria")
    if df_furos.empty:
        df_furos = pd.DataFrame(columns=["id", "sonda_id"])
    if df_prod.empty:
        df_prod = pd.DataFrame(columns=["id", "data", "sonda_id", "furo_id", "prof_inicial", "prof_final", "horas_trabalhadas", "horas_paradas", "motivo_parada"])
    if perfil_atual != "Admin":
        df_sondas = df_sondas[df_sondas["id"] == sonda_id_atual]
        df_furos = df_furos[df_furos["sonda_id"] == sonda_id_atual]
        df_prod = df_prod[df_prod["sonda_id"] == sonda_id_atual]
    if not df_prod.empty:
        df_prod = df_prod.merge(df_sondas[["id", "codigo"]], left_on="sonda_id", right_on="id", how="left", suffixes=("", "_sonda"))
        df_prod.rename(columns={"codigo": "sonda"}, inplace=True)
        df_prod["avanco"] = df_prod["prof_final"] - df_prod["prof_inicial"]
        df_prod = df_prod.sort_values("data", ascending=False)

    tab_hist, tab_novo = st.tabs(["Histórico", "Registrar Apontamento"])

    with tab_hist:
        if not df_prod.empty:
            st.dataframe(df_prod, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("🗑️ Excluir Registro de Apontamento")
            
            opcoes_prod = [f"ID: {row['id']} | Data: {row['data']} | Furo: {row['furo_id']} | Avanço: {row['avanco']}m" for _, row in df_prod.iterrows()]
            prod_excluir_str = st.selectbox("Selecione o Apontamento para excluir:", opcoes_prod)
            prod_id_excluir = int(prod_excluir_str.split(" | ")[0].replace("ID: ", ""))

            if st.button("Excluir Apontamento Selecionado", type="primary"):
                delete_row("producao_diaria", "id", prod_id_excluir)
                st.success("Apontamento excluído com sucesso!")
                st.rerun()
        else:
            st.info("Nenhum registro encontrado.")

    with tab_novo:
        if not df_sondas.empty and not df_furos.empty:
            with st.form("form_prod", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                dt = c1.date_input("Data", date.today())
                sonda_sel = c2.selectbox("Sonda", df_sondas["codigo"].tolist())
                furo_sel = c3.selectbox("Furo", df_furos["id"].tolist())

                c4, c5, c6 = st.columns(3)
                p_ini = c4.number_input("Prof. Inicial (m)", min_value=0.0, step=0.1)
                p_fin = c5.number_input("Prof. Final (m)", min_value=0.0, step=0.1)
                h_trab = c6.number_input("Horas Trabalhadas", min_value=0.0, step=0.5)

                c7, c8 = st.columns(2)
                h_par = c7.number_input("Horas Paradas", min_value=0.0, step=0.5)
                motivo = c8.text_input("Motivo Parada")

                if st.form_submit_button("Salvar Apontamento", type="primary"):
                    s_id = df_sondas[df_sondas["codigo"] == sonda_sel]["id"].values[0]
                    insert_row("producao_diaria", {
                        "data": str(dt),
                        "sonda_id": int(s_id),
                        "furo_id": furo_sel,
                        "prof_inicial": p_ini,
                        "prof_final": p_fin,
                        "horas_trabalhadas": h_trab,
                        "horas_paradas": h_par,
                        "motivo_parada": motivo,
                    })
                    st.success("Dados de produção salvos com sucesso!")
                    st.rerun()
        else:
            st.warning("É necessário possuir furos e sondas vinculados para registrar o apontamento.")

# ------------------------------------------------------------------------------
# 4. CONTROLE DE FUROS (COM CAPTURA DE GPS AUTOMÁTICO)
# ------------------------------------------------------------------------------
elif opcao == "Controle de Furos":
    botao_voltar_dashboard()
    st.title("Controle de Furos de Sondagem")
    st.markdown("---")

    df_sondas = table_df("sondas")
    df_furos = table_df("furos")
    if df_sondas.empty:
        df_sondas = pd.DataFrame(columns=["id", "codigo"])
    if df_furos.empty:
        df_furos = pd.DataFrame(columns=["id", "sonda_id", "coord_e", "coord_n", "cota", "prof_planejada", "prof_executada", "situacao"])
    if perfil_atual != "Admin":
        df_sondas = df_sondas[df_sondas["id"] == sonda_id_atual]
        df_furos = df_furos[df_furos["sonda_id"] == sonda_id_atual]
    if not df_furos.empty and not df_sondas.empty:
        df_furos = df_furos.merge(df_sondas[["id", "codigo"]], left_on="sonda_id", right_on="id", how="left", suffixes=("", "_sonda"))
        df_furos.rename(columns={"codigo": "sonda_codigo"}, inplace=True)

    tab_furos_list, tab_furos_novo = st.tabs(["Furos Cadastrados", "Novo Furo"])

    with tab_furos_list:
        if not df_furos.empty:
            st.dataframe(df_furos, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("🗑️ Excluir Furo")
            
            opcoes_furo = df_furos["id"].tolist()
            furo_excluir = st.selectbox("Selecione o Furo para excluir:", opcoes_furo)

            if st.button("Excluir Furo Selecionado", type="primary"):
                delete_row("furos", "id", furo_excluir)
                st.success(f"Furo '{furo_excluir}' excluído com sucesso!")
                st.rerun()
        else:
            st.info("Nenhum furo cadastrado.")

    with tab_furos_novo:
        if not df_sondas.empty:
            st.subheader("Captura de Localização via GPS")
            st.caption("Clique no botão abaixo para capturar as coordenadas exatas de onde você está no campo:")
            
            # Captura a localização atual via navegador/dispositivo
            location = streamlit_geolocation()

            lat_auto = location.get("latitude") if location else None
            lon_auto = location.get("longitude") if location else None
            alt_auto = location.get("altitude") if location else None

            if lat_auto and lon_auto:
                st.success(f"📍 Coordenadas capturadas: Lat {lat_auto:.6f}, Lon {lon_auto:.6f}")
            else:
                st.info("Clique no ícone de GPS acima para obter as coordenadas automaticamente ou preencha manualmente.")

            with st.form("form_furo", clear_on_submit=True):
                c1, c2 = st.columns(2)
                furo_id = c1.text_input("Identificação do Furo (ex: F-01)")
                sonda_sel = c2.selectbox("Sonda Responsável", df_sondas["codigo"].tolist())

                c3, c4, c5 = st.columns(3)
                coord_e = c3.number_input(
                    "Longitude / Easting", 
                    value=float(lon_auto) if lon_auto is not None else 0.0, 
                    format="%.6f"
                )
                coord_n = c4.number_input(
                    "Latitude / Northing", 
                    value=float(lat_auto) if lat_auto is not None else 0.0, 
                    format="%.6f"
                )
                cota = c5.number_input(
                    "Cota / Altitude (Z)", 
                    value=float(alt_auto) if alt_auto is not None else 0.0, 
                    format="%.2f"
                )

                c6, c7 = st.columns(2)
                prof_plan = c6.number_input("Prof. Planejada (m)", min_value=0.0, step=1.0)
                situacao = c7.selectbox("Situação", ["Planejado", "Em Andamento", "Concluído", "Cancelado"])

                if st.form_submit_button("Cadastrar Furo", type="primary"):
                    s_id = df_sondas[df_sondas["codigo"] == sonda_sel]["id"].values[0]
                    try:
                        insert_row("furos", {
                            "id": furo_id,
                            "sonda_id": int(s_id),
                            "coord_e": coord_e,
                            "coord_n": coord_n,
                            "cota": cota,
                            "prof_planejada": prof_plan,
                            "situacao": situacao,
                        })
                        st.success("Furo adicionado com sucesso!")
                        st.rerun()
                    except Exception as e:
                        if "duplicate" in str(e).lower() or "unique" in str(e).lower():
                            st.error("Identificação do Furo já existe.")
                        else:
                            st.error(f"Erro ao cadastrar furo: {e}")

# ------------------------------------------------------------------------------
# 5. BOLETIM GEOLÓGICO
# ------------------------------------------------------------------------------
elif opcao == "Boletim Geológico":
    botao_voltar_dashboard()
    st.title("Boletim Geológico de Campo")
    st.markdown("---")

    df_furos = table_df("furos")
    df_geo = table_df("boletim_geologico")
    if df_furos.empty:
        df_furos = pd.DataFrame(columns=["id", "sonda_id"])
    if df_geo.empty:
        df_geo = pd.DataFrame(columns=["id", "furo_id", "de_m", "ate_m", "recuperacao_m", "rqd_m", "litologia", "descricao_geologica", "n_amostra", "observacoes", "foto_url"])
    if perfil_atual != "Admin":
        df_furos = df_furos[df_furos["sonda_id"] == sonda_id_atual]
        df_geo = df_geo[df_geo["furo_id"].isin(df_furos["id"].tolist())]
    if not df_geo.empty:
        df_geo = df_geo.sort_values(["furo_id", "de_m"])

    tab_bg_hist, tab_bg_novo = st.tabs(["Registros Geológicos", "Registrar Intervalo"])

    with tab_bg_hist:
        if not df_geo.empty:
            st.dataframe(df_geo, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("🗑️ Excluir Registro do Boletim")
            
            lista_ids = df_geo["id"].tolist()
            id_para_excluir = st.selectbox("Selecione o ID do registro que deseja remover:", lista_ids)
            
            dados_registro = df_geo[df_geo["id"] == id_para_excluir].iloc[0]
            st.caption(f"Furo: **{dados_registro['furo_id']}** | Trecho: **{dados_registro['de_m']}m - {dados_registro['ate_m']}m** | Litologia: **{dados_registro['litologia']}**")
            
            if st.button("Excluir Registro Selecionado", type="primary"):
                delete_row("boletim_geologico", "id", int(id_para_excluir))

                st.success(f"Registro ID {id_para_excluir} excluído com sucesso!")
                st.rerun()
        else:
            st.info("Nenhum boletim registrado.")

    with tab_bg_novo:
        if not df_furos.empty:
            lista_furos = df_furos["id"].tolist()
            
            furo_sel = st.selectbox("Selecione o Furo para Registro", lista_furos)

            ultimo_resultado = (
                get_supabase()
                .table("boletim_geologico")
                .select("ate_m")
                .eq("furo_id", furo_sel)
                .order("ate_m", desc=True)
                .limit(1)
                .execute()
            )
            ultimo_ate = (_rows(ultimo_resultado)[0].get("ate_m") if _rows(ultimo_resultado) else None)

            de_auto = float(ultimo_ate) if ultimo_ate is not None else 0.0

            st.subheader("Intervalo e Métricas de Avanço")
            col_de, col_ate, col_avanco = st.columns(3)
            
            de_m = col_de.number_input("De (m)", min_value=0.0, value=de_auto, step=0.1, key="de_m_input")
            ate_m = col_ate.number_input("Até (m)", min_value=de_m, value=max(de_m, de_auto), step=0.1, key="ate_m_input")
            
            avanco_m = round(ate_m - de_m, 2)
            col_avanco.metric("Avanço Automático (m)", f"{avanco_m:.2f} m")

            with st.form("form_bg", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                rec_m = c1.number_input("Recuperação (m)", min_value=0.0, max_value=float(avanco_m) if avanco_m > 0 else 999.0, step=0.01)
                
                rec_pct = (rec_m / avanco_m * 100) if avanco_m > 0 else 0.0
                c2.text_input("Recuperação (%)", value=f"{rec_pct:.1f}%", disabled=True)
                
                rqd_m = c3.number_input("RQD (m)", min_value=0.0, max_value=float(avanco_m) if avanco_m > 0 else 999.0, step=0.01)

                c4, c5 = st.columns(2)
                litologia = c4.text_input("Litologia")
                n_amostra = c5.text_input("Nº Amostra")

                foto = st.file_uploader("Foto da Caixa de Testemunho", type=["jpg", "png", "jpeg"])
                desc = st.text_area("Descrição Geológica")
                obs = st.text_area("Observações")

                if st.form_submit_button("Salvar Boletim", type="primary"):
                    if ate_m <= de_m and avanco_m == 0:
                        st.error("A profundidade final 'Até (m)' deve ser maior que a profundidade inicial 'De (m)'.")
                    else:
                        foto_url = upload_foto_supabase(foto, foto.name) if foto else ""
                        insert_row("boletim_geologico", {
                            "furo_id": furo_sel,
                            "de_m": de_m,
                            "ate_m": ate_m,
                            "recuperacao_m": rec_m,
                            "rqd_m": rqd_m,
                            "litologia": litologia,
                            "descricao_geologica": desc,
                            "n_amostra": n_amostra,
                            "observacoes": obs,
                            "foto_url": foto_url,
                        })

                        # O registro é salvo apenas uma vez no Supabase.
                        # Mantemos a função de upload da foto separada.

                        st.success("Boletim Geológico salvo com sucesso!")
                        st.rerun()
        else:
            st.warning("Nenhum furo cadastrado para registrar boletim.")

# ------------------------------------------------------------------------------
# 6. GESTÃO DE USUÁRIOS (EXCLUSIVO ADMIN)
# ------------------------------------------------------------------------------
elif opcao == "Gestão de Usuários" and perfil_atual == "Admin":
    botao_voltar_dashboard()
    st.title("Vinculação de Usuários e Sondas")
    st.markdown("---")

    df_sondas = table_df("sondas")
    df_users = table_df("usuarios")
    if df_sondas.empty:
        df_sondas = pd.DataFrame(columns=["id", "codigo"])
    if df_users.empty:
        df_users = pd.DataFrame(columns=["id", "usuario", "perfil", "sonda_id"])
    if not df_users.empty and not df_sondas.empty:
        df_users = df_users.merge(df_sondas[["id", "codigo"]], left_on="sonda_id", right_on="id", how="left", suffixes=("", "_sonda"))
        df_users.rename(columns={"codigo": "sonda_vinculada"}, inplace=True)

    tab_u_lista, tab_u_novo = st.tabs(["Usuários Cadastrados", "Novo Usuário"])

    with tab_u_lista:
        if not df_users.empty:
            st.dataframe(df_users, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("🗑️ Excluir Usuário")
            
            # Não permite excluir o usuário admin principal para evitar bloqueio
            users_filtrados = df_users[df_users["usuario"] != "admin"]
            
            if not users_filtrados.empty:
                opcoes_user = [f"ID: {row['id']} - {row['usuario']} ({row['perfil']})" for _, row in users_filtrados.iterrows()]
                user_excluir_str = st.selectbox("Selecione o Usuário para excluir:", opcoes_user)
                user_id_excluir = int(user_excluir_str.split(" - ")[0].replace("ID: ", ""))

                if st.button("Excluir Usuário Selecionado", type="primary"):
                    delete_row("usuarios", "id", user_id_excluir)
                    st.success("Usuário excluído com sucesso!")
                    st.rerun()
            else:
                st.caption("Apenas o usuário 'admin' principal está cadastrado (protegido contra exclusão).")
        else:
            st.info("Nenhum usuário cadastrado.")

    with tab_u_novo:
        with st.form("form_user", clear_on_submit=True):
            new_user = st.text_input("Nome do Usuário")
            new_pass = st.text_input("Senha", type="password")
            new_perfil = st.selectbox("Perfil", ["Geólogo", "Operador", "Admin"])
            
            sonda_opcoes = ["Nenhuma / Admin"] + (df_sondas["codigo"].tolist() if not df_sondas.empty else [])
            new_sonda = st.selectbox("Sonda Vinculada", sonda_opcoes)

            if st.form_submit_button("Cadastrar Usuário", type="primary"):
                if new_user and new_pass:
                    s_id = None
                    if new_sonda != "Nenhuma / Admin" and not df_sondas.empty:
                        s_id = int(df_sondas[df_sondas["codigo"] == new_sonda]["id"].values[0])

                    try:
                        insert_row("usuarios", {
                            "usuario": new_user,
                            "senha": hash_senha(new_pass),
                            "perfil": new_perfil,
                            "sonda_id": s_id,
                        })
                        st.success("Usuário cadastrado com sucesso!")
                        st.rerun()
                    except Exception as e:
                        if "duplicate" in str(e).lower() or "unique" in str(e).lower():
                            st.error("Usuário já existe.")
                        else:
                            st.error(f"Erro ao cadastrar usuário: {e}")
